import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import time
import re

BASE_URL = "https://www.barcin.com/erkek-kosu-ayakkabisi/?attributes_integration_brand=adidas&attributes_integration_gender=Erkek"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "tr-TR,tr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def get_total_pages(soup):
    for text in soup.find_all(string=re.compile(r'\d+\s*[Üü]r[üu]n')):
        m = re.search(r'(\d+)\s*[Üü]r[üu]n', text)
        if m:
            total = int(m.group(1))
            print(f"Toplam ürün: {total}")
            return (total + 23) // 24
    return 1


def scrape_page(url):
    products = []
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ❌ {url} → {e}")
        return products

    soup = BeautifulSoup(resp.text, "html.parser")

    # Her ürün: div.product-info içinde h3 > a (ürün adı) ve span[data-testid="product-price"] (fiyat)
    for info_div in soup.find_all("div", class_="product-info"):
        # Ürün adı ve linki
        h3 = info_div.find("h3")
        if not h3:
            continue
        a = h3.find("a")
        if not a:
            continue

        name = a.get_text(strip=True)
        href = a.get("href", "")
        full_url = "https://www.barcin.com" + href if href.startswith("/") else href

        # Fiyatlar: span[data-testid="product-price"] — indirimli ürünlerde 2 tane olabilir
        price_spans = info_div.find_all("span", attrs={"data-testid": "product-price"})

        if len(price_spans) == 0:
            sale_price, original_price = "", ""
        elif len(price_spans) == 1:
            sale_price = price_spans[0].get_text(strip=True)
            original_price = ""
        else:
            # İlk span indirimli fiyat, ikinci normal fiyat
            sale_price = price_spans[0].get_text(strip=True)
            original_price = price_spans[1].get_text(strip=True)

        # İndirim oranı (varsa)
        discount = ""
        discount_el = info_div.find(string=re.compile(r'-%\d+|-\d+%'))
        if not discount_el:
            # Üst kart elementini de kontrol et
            parent_card = info_div.parent
            if parent_card:
                d = parent_card.find(string=re.compile(r'-\d+%'))
                if d:
                    discount = d.strip()

        if name:
            products.append({
                "name": name,
                "sale_price": sale_price,
                "original_price": original_price,
                "url": full_url
            })

    return products


def save_to_excel(products, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barcın"

    headers = ["#", "Ürün Adı", "Satış Fiyatı", "Normal Fiyat", "URL"]
    widths = [5, 65, 18, 18, 70]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 25

    for i, p in enumerate(products, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p["name"])
        ws.cell(row=row, column=3, value=p["sale_price"])
        ws.cell(row=row, column=4, value=p["original_price"])
        ws.cell(row=row, column=5, value=p["url"])
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                )

    ws.cell(row=len(products) + 3, column=1, value=f"Toplam: {len(products)} ürün")
    ws.cell(row=len(products) + 4, column=1, value=f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    wb.save(filename)
    print(f"📊 Excel kaydedildi: {filename}")


def main():
    print(f"🔍 Taranıyor: {BASE_URL}")

    resp = requests.get(BASE_URL, headers=HEADERS, timeout=15)
    soup = BeautifulSoup(resp.text, "html.parser")
    total_pages = get_total_pages(soup)
    print(f"Toplam sayfa: {total_pages}")

    all_products = []

    for page in range(1, total_pages + 1):
        if page == 1:
            url = BASE_URL
        elif "?" in BASE_URL:
            url = f"{BASE_URL}&page={page}"
        else:
            url = f"{BASE_URL}?page={page}"
        print(f"  Sayfa {page}/{total_pages}: {url}")
        products = scrape_page(url)
        print(f"    → {len(products)} ürün", end="")
        if products:
            print(f" | Örnek: {products[0]['name']} → {products[0]['sale_price']}")
        else:
            print()
        all_products.extend(products)
        time.sleep(0.8)

    # Duplicate temizle
    seen = set()
    unique = []
    for p in all_products:
        if p["url"] not in seen:
            seen.add(p["url"])
            unique.append(p)

    print(f"\n✅ Toplam {len(unique)} ürün bulundu.")

    if unique:
        fname = f"barcin_kadin_kosu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_to_excel(unique, fname)
    else:
        print("❌ Ürün bulunamadı!")


if __name__ == "__main__":
    main()