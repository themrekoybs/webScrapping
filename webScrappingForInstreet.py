import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import time
import re

URLS = [
    "https://www.instreet.com.tr/kosu-ayakkabisi?marka=adidas&cinsiyet=erkek",
    # "https://www.instreet.com.tr/marka/adidas?cinsiyet=kadin",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "tr-TR,tr;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def get_all_pages(base_url):
    urls = [base_url]
    try:
        resp = requests.get(base_url, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")

        total_products = 0
        for el in soup.find_all(string=re.compile(r'\d+\s*[Üü]r[üu]n')):
            m = re.search(r'(\d+)\s*[Üü]r[üu]n', el)
            if m:
                total_products = int(m.group(1))
                print(f"Toplam ürün: {total_products}")
                break

        page_nums = set()
        for a in soup.select("a[href*='page=']"):
            m = re.search(r'[?&]page=(\d+)', a.get("href", ""))
            if m:
                page_nums.add(int(m.group(1)))

        if not page_nums and total_products > 24:
            pages = (total_products + 23) // 24
            for pg in range(2, pages + 1):
                page_nums.add(pg)

        sep = "&" if "?" in base_url else "?"
        for pg in sorted(page_nums):
            urls.append(f"{base_url}{sep}page={pg}")

    except Exception as e:
        print(f"Sayfa listesi alınamadı: {e}")

    return urls


def scrape_page(url):
    products = []
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ❌ {url} → {e}")
        return products

    soup = BeautifulSoup(resp.text, "html.parser")

    # Ürün adı: class="product-name__wrapper" içindeki <a> linkleri
    name_links = soup.find_all("a", href=re.compile(r'^/urun/'))

    # Her linke karşılık gelen fiyatı bul
    # Strateji: aynı href'e sahip 3 farklı <a> var.
    # Fiyat: div.product-pricing-one__price
    # Önce tüm fiyatları sırayla topla
    price_divs = soup.find_all("div", class_="product-pricing-one__price")

    # Ürün adı linkleri: sadece h3.product-name__wrapper içerenleri al
    product_links = []
    seen_hrefs = set()
    for a in name_links:
        if a.find("h3", class_="product-name__wrapper") and a["href"] not in seen_hrefs:
            seen_hrefs.add(a["href"])
            product_links.append(a)

    for i, a in enumerate(product_links):
        href = a.get("href", "")

        # Marka
        brand = ""
        brand_el = a.find("span", class_="product__brand")
        if brand_el:
            brand = brand_el.get_text(strip=True)

        # Ürün adı
        name = ""
        name_el = a.find("span", class_="product__name")
        if name_el:
            name = name_el.get_text(strip=True)
        full_name = f"{brand} {name}".strip()

        # Fiyat: sıradaki fiyat div'i (index eşleşmesi)
        price = ""
        if i < len(price_divs):
            price = price_divs[i].get_text(strip=True)

        full_url = "https://www.instreet.com.tr" + href

        if full_name:
            products.append({
                "name": full_name,
                "price": price,
                "url": full_url
            })

    return products


def save_to_excel(products, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Instreet"

    headers = ["#", "Ürün Adı", "Fiyat", "URL"]
    widths = [5, 60, 20, 70]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 25

    for i, p in enumerate(products, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p["name"])
        ws.cell(row=row, column=3, value=p["price"])
        ws.cell(row=row, column=4, value=p["url"])
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
                )

    ws.cell(row=len(products) + 3, column=1, value=f"Toplam: {len(products)} ürün")
    ws.cell(row=len(products) + 4, column=1, value=f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    wb.save(filename)
    print(f"📊 Excel kaydedildi: {filename}")


def main():
    all_products = []

    for base_url in URLS:
        print(f"\n🔍 Taranıyor: {base_url}")
        pages = get_all_pages(base_url)
        print(f"  {len(pages)} sayfa bulundu.")

        for i, page_url in enumerate(pages, 1):
            print(f"  Sayfa {i}/{len(pages)}: {page_url}")
            products = scrape_page(page_url)
            print(f"    → {len(products)} ürün (örnek: {products[0] if products else 'yok'})")
            all_products.extend(products)
            time.sleep(0.8)

    # Duplicate temizle
    seen = set()
    unique = []
    for p in all_products:
        if p["url"] not in seen:
            seen.add(p["url"])
            unique.append(p)

    print(f"\n✅ Toplam {len(unique)} ürün bulundu.")

    if unique:
        fname = f"instreet_adidas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_to_excel(unique, fname)
    else:
        print("❌ Ürün bulunamadı!")


if __name__ == "__main__":
    main()