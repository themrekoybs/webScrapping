import asyncio
import re
import subprocess
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

URL = "https://www.instreet.com.tr/marka/adidas?gad_source=1&gad_campaignid=1718368539&gbraid=0AAAAADRqLqYwf76o32yqEFbpFm6cQYp4P&gclid=Cj0KCQiA5I_NBhDVARIsAOrqIsbLsBzTgxHgz-93JWu9-JzUtSZe1N0Epzp3ltEfzOXzDa2ufE7_Mj0aAiN2EALw_wcB&cinsiyet=kadin"

# Mac'te Chrome'un gerçek yolu
CHROME_PATH = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

async def scrape_adidas():
    products = []

    async with async_playwright() as p:
        # Sisteme kurulu Chrome'u kullan (bot tespitini atlatır)
        browser = await p.chromium.launch_persistent_context(
            user_data_dir="/tmp/adidas_chrome_profile",  # Geçici profil klasörü
            executable_path=CHROME_PATH,
            headless=False,  # Görünür modda çalıştır
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--start-maximized",
            ],
            viewport=None,  # Tam ekran için None
            locale="tr-TR",
            timezone_id="Europe/Istanbul",
        )

        page = await browser.new_page()

        # Webdriver özelliğini gizle
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = { runtime: {} };
        """)

        print("Sayfa yükleniyor...")
        await page.goto(URL, wait_until="domcontentloaded", timeout=90000)
        await page.wait_for_timeout(4000)

        # Cookie/popup kapat
        popup_selectors = [
            "[data-auto-id='glass-gdpr-default-consent-accept-button']",
            "button:has-text('Kabul Et')",
            "button:has-text('Accept All')",
            "button:has-text('Tümünü Kabul Et')",
            "#onetrust-accept-btn-handler",
        ]
        for sel in popup_selectors:
            try:
                btn = page.locator(sel)
                if await btn.count() > 0:
                    await btn.first.click()
                    print(f"Popup kapatıldı.")
                    await page.wait_for_timeout(2000)
                    break
            except:
                pass

        # Sayfada hangi selector'lar çalışıyor bul
        print("Selector taranıyor...")
        product_selector = None
        test_selectors = [
            "[data-auto-id='glass-product-card']",
            "article[class*='product']",
            "[class*='product-card']",
            "[class*='ProductCard']",
            "[class*='plp'] article",
            "li[class*='product']",
            "article",
        ]
        for sel in test_selectors:
            count = await page.locator(sel).count()
            if count > 3:
                print(f"✅ Selector bulundu: '{sel}' → {count} ürün")
                product_selector = sel
                break

        if not product_selector:
            print("❌ Ürün kartı bulunamadı! HTML kaydediliyor...")
            html = await page.content()
            with open("adidas_debug.html", "w", encoding="utf-8") as f:
                f.write(html)
            print("adidas_debug.html kaydedildi. Bu dosyayı paylaşın.")
            await browser.close()
            return

        # Scroll yaparak tüm ürünleri yükle
        print("Ürünler yükleniyor...")
        prev_count = 0
        no_change = 0

        while no_change < 5:
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(2500)

            # "Daha Fazla Göster" butonu
            for btn_text in ["Daha Fazla Göster", "Daha fazla göster", "Load More", "Show More"]:
                try:
                    btn = page.locator(f"button:has-text('{btn_text}')")
                    if await btn.count() > 0 and await btn.first.is_visible():
                        await btn.first.click()
                        print(f"  '{btn_text}' tıklandı.")
                        await page.wait_for_timeout(3000)
                        break
                except:
                    pass

            count = await page.locator(product_selector).count()
            print(f"  Ürün sayısı: {count}")

            if count == prev_count:
                no_change += 1
            else:
                no_change = 0
                prev_count = count

        print(f"\n{prev_count} ürün bulundu. Veriler çekiliyor...")

        # Ürün bilgilerini çek
        cards = page.locator(product_selector)
        total = await cards.count()

        for i in range(total):
            try:
                card = cards.nth(i)

                # Ad
                name = ""
                for sel in ["[data-auto-id='glass-product-card-title']", "h3", "h2", "[class*='title']"]:
                    el = card.locator(sel).first
                    if await el.count() > 0:
                        t = (await el.inner_text()).strip()
                        if t:
                            name = t
                            break

                # Alt başlık
                subtitle = ""
                for sel in ["[data-auto-id='glass-product-card-subtitle']", "[class*='subtitle']", "[class*='category']"]:
                    el = card.locator(sel).first
                    if await el.count() > 0:
                        t = (await el.inner_text()).strip()
                        if t:
                            subtitle = t
                            break

                # Fiyat
                price = ""
                for sel in ["[data-auto-id='glass-product-card-price']", "[class*='price']"]:
                    el = card.locator(sel).first
                    if await el.count() > 0:
                        t = (await el.inner_text()).strip()
                        t = re.sub(r'\s+', ' ', t)
                        if t and any(c.isdigit() for c in t):
                            price = t
                            break

                # URL
                link = ""
                try:
                    href = await card.locator("a").first.get_attribute("href")
                    if href:
                        link = href if href.startswith("http") else "https://www.adidas.com.tr" + href
                except:
                    pass

                if name:
                    products.append({"name": name, "subtitle": subtitle, "price": price, "url": link})
                    if i % 20 == 0:
                        print(f"  {i}/{total} işlendi...")

            except Exception as e:
                continue

        await browser.close()

    print(f"\n✅ {len(products)} ürün çekildi.")
    if products:
        save_to_excel(products)
    else:
        print("❌ Ürün verisi boş!")


def save_to_excel(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adidas Koşu Ayakkabıları"

    # Header
    headers = ["#", "Ürün Adı", "Alt Başlık / Kategori", "Fiyat", "URL"]
    widths   = [5,   45,         30,                       20,      60]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for i, p in enumerate(products, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=p["name"])
        ws.cell(row=row, column=3, value=p["subtitle"])
        ws.cell(row=row, column=4, value=p["price"])
        ws.cell(row=row, column=5, value=p["url"])
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    ws.cell(row=len(products)+3, column=1, value=f"Toplam: {len(products)} ürün")
    ws.cell(row=len(products)+4, column=1, value=f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}")

    fname = f"adidas_kosu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    print(f"📊 Excel kaydedildi: {fname}")


if __name__ == "__main__":
    asyncio.run(scrape_adidas())